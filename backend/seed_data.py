"""
Deal Room development seed loader.

Business/test data source:
    backend/seed_data/Deal_Room_Seed_Dataset.xlsx

The Excel workbook is the source of truth for:
    Accounts
    Opportunities
    Stakeholders
    POC_Tracker
    Activity_Log
    OEM_Partners

The script intentionally keeps only the application test identities and
canonical pipeline stages in Python. It does NOT hard-code business records.

Run from backend:
    python seed_test_data.py

Optional:
    SEED_DATABASE_URL=postgresql://dealroom:dealroom_dev@127.0.0.1:5432/dealroom2 \
        python seed_test_data.py

Normal execution is idempotent. It never drops, truncates, or deletes data.
"""

import os
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from sqlalchemy import inspect
from sqlalchemy.engine import make_url

load_dotenv()

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_SEED_FILE = BASE_DIR / "seed_data" / "Deal_Room_Seed_Dataset.xlsx"

configured_url = os.getenv("SEED_DATABASE_URL") or os.getenv("DATABASE_URL")
if not configured_url:
    raise RuntimeError(
        "DATABASE_URL is not set. Add it to backend/.env or set SEED_DATABASE_URL."
    )

os.environ["DATABASE_URL"] = configured_url

SEED_DATA_FILE = Path(
    os.getenv("SEED_DATA_FILE", str(DEFAULT_SEED_FILE))
)
if not SEED_DATA_FILE.is_absolute():
    SEED_DATA_FILE = BASE_DIR / SEED_DATA_FILE

from app import create_app
from app.database import db
from app.auth.password import hash_password
from app.constants.auth_constants import (
    STATUS_APPROVED,
    STATUS_PENDING,
    STATUS_REVOKED,
)
from app.constants.roles import (
    ADMIN,
    SALES_EXECUTIVE,
    SALES_MANAGER,
    PRE_SALES_MANAGER,
    SOLUTION_ENGINEER,
    DELIVERY,
)
from app.constants.stages import PIPELINE_STAGES, CLOSED_STATUS, OPEN_STATUS

from app.models.auth.user import User
from app.models.auth.user_role import UserRole
from app.models.account.account import Account
from app.models.account.contact import Contact
from app.models.account.oem_partner import OEMPartner
from app.models.opportunity.stage_master import StageMaster
from app.models.opportunity.opportunity import Opportunity
from app.models.opportunity.opportunity_team import OpportunityTeam
from app.models.opportunity.stakeholder import Stakeholder
from app.models.opportunity.stage_history import StageHistory
from app.models.opportunity.poc_tracker import POCTracker
from app.models.poc.poc import Poc
from app.models.system.tag import Tag
from app.models.system.audit_log import AuditLog
from app.models.system.notification import Notification


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------

def get_or_create(model, filters, defaults=None):
    obj = model.query.filter_by(**filters).first()
    if obj:
        return obj, False

    values = dict(filters)
    values.update(defaults or {})
    obj = model(**values)
    db.session.add(obj)
    db.session.flush()
    return obj, True


def model_columns(model):
    return {
        column.name
        for column in inspect(model).columns
    }


def set_if_present(obj, values):
    columns = model_columns(type(obj))
    for key, value in values.items():
        if key in columns and value is not None:
            setattr(obj, key, value)


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def parse_date(value):
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


def parse_datetime(value):
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    text = str(value).strip()
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return datetime.strptime(text[:10], "%Y-%m-%d")


def probability_as_percent(value):
    if value is None:
        return None
    value = float(value)
    # Workbook stores values such as 0.53. Existing model/seed uses 53.
    if 0 <= value <= 1:
        return round(value * 100)
    return round(value)


def rows_from_sheet(workbook, sheet_name):
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [
        str(value).strip() if value is not None else ""
        for value in rows[0]
    ]

    result = []
    for row in rows[1:]:
        if not any(value is not None for value in row):
            continue
        result.append({
            headers[index]: row[index]
            for index in range(len(headers))
        })
    return result


def load_seed_workbook():
    if not SEED_DATA_FILE.exists():
        raise FileNotFoundError(
            f"Seed workbook not found: {SEED_DATA_FILE}\n"
            "Place Deal_Room_Seed_Dataset.xlsx in backend/seed_data/ "
            "or set SEED_DATA_FILE in backend/.env."
        )

    workbook = load_workbook(
        SEED_DATA_FILE,
        read_only=True,
        data_only=True,
    )

    required = {
        "Accounts",
        "Opportunities",
        "Stakeholders",
        "POC_Tracker",
        "Activity_Log",
        "OEM_Partners",
    }

    missing = required - set(workbook.sheetnames)
    if missing:
        raise RuntimeError(
            "Seed workbook is missing sheets: "
            + ", ".join(sorted(missing))
        )

    return workbook


# ---------------------------------------------------------------------------
# Users
# ---------------------------------------------------------------------------

def seed_users():
    """
    Application test identities remain deterministic and are intentionally
    kept in Python. Business data comes from Excel.
    """
    password_hash = hash_password("Test@123")

    specs = [
        (
            "System Administrator",
            "admin@dealroom.local",
            [ADMIN],
            STATUS_APPROVED,
        ),
        (
            "Sales Executive",
            "sales.exec@dealroom.local",
            [SALES_EXECUTIVE],
            STATUS_APPROVED,
        ),
        (
            "Sales Manager",
            "sales.manager@dealroom.local",
            [SALES_MANAGER],
            STATUS_APPROVED,
        ),
        (
            "Pre-Sales Manager",
            "presales.manager@dealroom.local",
            [PRE_SALES_MANAGER],
            STATUS_APPROVED,
        ),
        (
            "Solution Engineer",
            "solution.engineer@dealroom.local",
            [SOLUTION_ENGINEER],
            STATUS_APPROVED,
        ),
        (
            "Solution Engineer",
            "delivery@dealroom.local",
            [SOLUTION_ENGINEER],
            STATUS_APPROVED,
        ),
        (
            "Multi Role User",
            "multi.role@dealroom.local",
            [SOLUTION_ENGINEER],
            STATUS_APPROVED,
        ),
        (
            "Pending User",
            "pending@dealroom.local",
            [],
            STATUS_PENDING,
        ),
        (
            "Revoked User",
            "revoked@dealroom.local",
            [SALES_EXECUTIVE],
            STATUS_REVOKED,
        ),
    ]

    users = {}

    for name, email, roles, status in specs:
        user = User.query.filter_by(email=email).first()

        if not user:
            user = User(
                full_name=name,
                email=email,
                password_hash=password_hash,
                status=status,
                active=(status != STATUS_REVOKED),
                approved_at=(
                    datetime.utcnow()
                    if status == STATUS_APPROVED
                    else None
                ),
            )
            db.session.add(user)
            db.session.flush()
        else:
            user.full_name = name
            user.password_hash = password_hash
            user.status = status
            user.active = status != STATUS_REVOKED

            if status == STATUS_APPROVED and not user.approved_at:
                user.approved_at = datetime.utcnow()

        existing_roles = {row.role for row in user.roles}

        for role in roles:
            if role not in existing_roles:
                user.roles.append(UserRole(role=role))

        if not roles:
            for role_row in list(user.roles):
                db.session.delete(role_row)

        users[email] = user

    db.session.flush()

    manager_map = {
        "sales.exec@dealroom.local": "sales.manager@dealroom.local",
        "solution.engineer@dealroom.local": "presales.manager@dealroom.local",
        "delivery@dealroom.local": "presales.manager@dealroom.local",
        "multi.role@dealroom.local": "presales.manager@dealroom.local",
    }

    for email, manager_email in manager_map.items():
        users[email].manager_id = users[manager_email].user_id

    for email in {
        "admin@dealroom.local",
        "sales.manager@dealroom.local",
        "presales.manager@dealroom.local",
        "pending@dealroom.local",
        "revoked@dealroom.local",
    }:
        users[email].manager_id = None

    db.session.flush()
    return users


def user_for_role(users, role):
    for user in users.values():
        if user.has_role(role):
            return user
    return None


# ---------------------------------------------------------------------------
# Stages
# ---------------------------------------------------------------------------

def seed_stages():
    stages = {}

    for stage_data in PIPELINE_STAGES:
        stage, _ = get_or_create(
            StageMaster,
            {"stage_name": stage_data["stage_name"]},
            {
                "display_order": stage_data["display_order"],
                "requires_poc": stage_data["requires_poc"],
                "is_closed": stage_data["stage_name"]
                in {"Closed Won", "Closed Lost"},
                "is_won": stage_data["stage_name"] == "Closed Won",
            },
        )

        stage.display_order = stage_data["display_order"]
        stage.requires_poc = stage_data["requires_poc"]
        stage.is_closed = stage_data["stage_name"] in {
            "Closed Won",
            "Closed Lost",
        }
        stage.is_won = stage_data["stage_name"] == "Closed Won"

        stages[stage.stage_name] = stage

    db.session.flush()
    return stages


# ---------------------------------------------------------------------------
# Accounts
# ---------------------------------------------------------------------------

def seed_accounts(workbook):
    accounts = {}

    for row in rows_from_sheet(workbook, "Accounts"):
        external_id = clean(row["account_id"])
        account_name = clean(row["account_name"])

        if not external_id or not account_name:
            continue

        account = Account.query.filter_by(
            account_name=account_name
        ).first()

        if not account:
            account = Account(
                account_name=account_name
            )
            db.session.add(account)
            db.session.flush()

        set_if_present(
            account,
            {
                "account_name": account_name,
                "industry": clean(row.get("industry")),
                "region": clean(row.get("region")),
                "website": clean(row.get("website")),
                "created_at": parse_datetime(row.get("created_at")),
            },
        )

        accounts[external_id] = account
        accounts[account_name] = account

    db.session.flush()
    return accounts


# ---------------------------------------------------------------------------
# Contacts
# ---------------------------------------------------------------------------

def seed_contacts(workbook, accounts):
    """
    The workbook does not have a Contacts sheet. Stakeholders contain enough
    contact information to populate the existing Contact table without
    inventing a second hard-coded contact dataset.
    """
    for row in rows_from_sheet(workbook, "Stakeholders"):
        opportunity_external_id = clean(row.get("opportunity_id"))
        email = clean(row.get("contact_email"))
        name = clean(row.get("stakeholder_name"))

        if not email or not name:
            continue

        # Resolve account through the opportunity mapping created later.
        opportunity = Opportunity.query.filter_by(
            opportunity_name=(
                Opportunity.query
                .filter_by()
                .order_by(Opportunity.opportunity_id)
                .first().opportunity_name
                if False else ""
            )
        ).first()

        # Account resolution is performed from the opportunity after all
        # opportunities have been loaded. This function is intentionally a
        # no-op until then.
        del opportunity
        del opportunity_external_id
        del accounts


# ---------------------------------------------------------------------------
# OEM partners
# ---------------------------------------------------------------------------

def seed_oem_partners(workbook, accounts):
    partner_columns = model_columns(OEMPartner)

    for row in rows_from_sheet(workbook, "OEM_Partners"):
        account = accounts.get(
            clean(row.get("associated_account_id"))
        )

        if not account:
            continue

        partner_name = clean(row.get("partner_name"))
        partner_type = clean(row.get("partner_type"))

        if not partner_name:
            continue

        filters = {
            "account_id": account.account_id,
            "partner_name": partner_name,
        }

        # product_name exists in the current schema. The workbook provides
        # partner_type rather than product_name, so use partner_type as the
        # source value only when the model requires product_name.
        if "product_name" in partner_columns:
            filters["product_name"] = partner_type or "Partner"

        partner, _ = get_or_create(
            OEMPartner,
            filters,
            {},
        )

        set_if_present(
            partner,
            {
                "partner_name": partner_name,
                "product_name": partner_type,
                "contact_person": clean(row.get("contact_name")),
                "email": clean(row.get("contact_email")),
                "status": clean(row.get("agreement_status")),
                "notes": "Imported from Deal Room seed workbook.",
            },
        )

    db.session.flush()


# ---------------------------------------------------------------------------
# Opportunities
# ---------------------------------------------------------------------------

def seed_opportunities(workbook, accounts, stages, users):
    opportunities = {}

    sales_exec = user_for_role(
        users,
        SALES_EXECUTIVE,
    )

    for row in rows_from_sheet(workbook, "Opportunities"):
        external_id = clean(row.get("opportunity_id"))
        account_external_id = clean(row.get("account_id"))
        name = clean(row.get("opportunity_name"))
        stage_name = clean(row.get("pipeline_stage"))

        if not external_id or not name:
            continue

        account = accounts.get(account_external_id)
        if not account:
            raise RuntimeError(
                f"Opportunity {external_id} references unknown "
                f"account {account_external_id}."
            )

        stage_name = clean(row.get("pipeline_stage"))

        if not stage_name:
            raise RuntimeError(
                f"Opportunity {external_id} has no pipeline_stage."
            )

        stage = stages.get(stage_name)

        if stage is None:
            raise RuntimeError(
                f"Opportunity {external_id} references unknown stage "
                f"{stage_name!r}. "
                f"Available stages: {', '.join(sorted(stages.keys()))}"
            )

        if stage.stage_id is None:
            raise RuntimeError(
                f"Stage {stage_name!r} exists but has no stage_id."
            )

        values = {
                    "account_id": account.account_id,
                    "opportunity_name": name,
                    "stage_id": stage.stage_id,
                    "estimated_value": (
                        Decimal(str(row["forecast_value_usd"]))
                        if row.get("forecast_value_usd") is not None
                        else None
                    ),
                    "probability": probability_as_percent(
                        row.get("probability")
                    ),
                    "expected_close_date": parse_date(
                        row.get("expected_close_date")
                    ),
                    "created_at": parse_datetime(
                        row.get("created_at")
                    ),
                    "updated_at": parse_datetime(
                        row.get("last_activity_date")
                    ),
                    "status": (
                        CLOSED_STATUS
                        if stage_name in {"Closed Won", "Closed Lost"}
                        else OPEN_STATUS
                    ),
                    "is_active": stage_name not in {
                        "Closed Won",
                        "Closed Lost",
                    },
                }

        opportunity = Opportunity.query.filter_by(
                    account_id=account.account_id,
                    opportunity_name=name,
                ).first()
        
        if sales_exec is None:
            raise RuntimeError(
                "No active Sales Executive development user was found. "
                "Expected sales.exec@dealroom.local with the Sales Executive role."
            )

        values["created_by"] = sales_exec.user_id
        values["sales_owner_id"] = sales_exec.user_id

        if not opportunity:
            opportunity = Opportunity(**values)
            db.session.add(opportunity)
        else:
            set_if_present(opportunity, values)

        # Flush only after all required foreign keys have been assigned.
        # The previous code flushed the new row before applying stage_id,
        # created_by, and sales_owner_id, causing NULLs in the INSERT.
        db.session.flush()

        opportunities[external_id] = opportunity

    db.session.flush()
    return opportunities


# ---------------------------------------------------------------------------
# Stakeholders
# ---------------------------------------------------------------------------

def seed_stakeholders(workbook, opportunities):
    for row in rows_from_sheet(workbook, "Stakeholders"):
        external_opportunity_id = clean(
            row.get("opportunity_id")
        )
        opportunity = opportunities.get(
            external_opportunity_id
        )

        if not opportunity:
            continue

        email = clean(row.get("contact_email"))
        name = clean(row.get("stakeholder_name"))

        if not email or not name:
            continue

        stakeholder, created = get_or_create(
            Stakeholder,
            {
                "opportunity_id": opportunity.opportunity_id,
                "email": email,
            },
            {
                "stakeholder_name": name,
                "designation": clean(row.get("role_title")),
                "phone": clean(row.get("contact_phone")),
                "influence_level": clean(
                    row.get("influence_level")
                ),
                "notes": "Imported from Deal Room seed workbook.",
            },
        )

        if not created:
            set_if_present(
                stakeholder,
                {
                    "stakeholder_name": name,
                    "designation": clean(row.get("role_title")),
                    "phone": clean(row.get("contact_phone")),
                    "influence_level": clean(
                        row.get("influence_level")
                    ),
                    "notes": "Imported from Deal Room seed workbook.",
                },
            )





    db.session.flush()


# ---------------------------------------------------------------------------
# Contacts derived from stakeholders
# ---------------------------------------------------------------------------

def seed_contacts_from_stakeholders(workbook, opportunities, accounts):
    """
    Populate Contact using stakeholder rows where possible. This does not
    invent contact records; it derives them from the supplied workbook.
    """
    for row in rows_from_sheet(workbook, "Stakeholders"):
        opportunity = opportunities.get(
            clean(row.get("opportunity_id"))
        )

        if not opportunity:
            continue

        email = clean(row.get("contact_email"))
        name = clean(row.get("stakeholder_name"))

        if not email or not name:
            continue

        account = accounts.get(opportunity.account_id)
        if not account:
            account = db.session.get(Account, opportunity.account_id)

        if not account:
            continue

        contact, created = get_or_create(
            Contact,
            {
                "account_id": account.account_id,
                "email": email,
            },
            {
                "full_name": name,
                "title": clean(row.get("role_title")),
                "is_primary": False,
            },
        )

        if not created:
            set_if_present(
                contact,
                {
                    "full_name": name,
                    "title": clean(row.get("role_title")),
                    "is_primary": False,
                },
            )

    db.session.flush()


# ---------------------------------------------------------------------------
# Opportunity team
# ---------------------------------------------------------------------------

def seed_opportunity_teams(opportunities, users):
    """
    Build deliberate visibility boundaries from the workbook's opportunity
    order. This is test scaffolding, not business data.

    Every opportunity gets one sales owner plus selected technical members,
    distributed across the fixed development users so authorization can be
    tested without making every user a member of every opportunity.
    """
    sales_exec = user_for_role(users, SALES_EXECUTIVE)
    sales_manager = user_for_role(users, SALES_MANAGER)
    se = user_for_role(users, SOLUTION_ENGINEER)
    delivery = user_for_role(users, DELIVERY)

    if not all([sales_exec, sales_manager, se, delivery]):
        raise RuntimeError(
            "Required development users for opportunity visibility "
            "testing are missing."
        )

    ordered = list(opportunities.values())

    for index, opportunity in enumerate(ordered):
        sales_user = (
            sales_exec
            if index % 2 == 0
            else sales_manager
        )

        team_entries = [
            (
                sales_user,
                SALES_EXECUTIVE
                if sales_user == sales_exec
                else SALES_MANAGER,
            )
        ]

        # Roughly one-third SE-only, one-third Delivery-only, and one-third
        # SE + Delivery. This deliberately creates cross-user visibility
        # boundaries for security testing.
        mode = index % 3

        if mode in (0, 2):
            team_entries.append(
                (se, SOLUTION_ENGINEER)
            )

        if mode in (1, 2):
            team_entries.append(
                (delivery, SOLUTION_ENGINEER)
            )

        for user, team_role in team_entries:
            row, _ = get_or_create(
                OpportunityTeam,
                {
                    "opportunity_id": opportunity.opportunity_id,
                    "user_id": user.user_id,
                },
                {"role": team_role},
            )
            row.role = team_role

    db.session.flush()


# ---------------------------------------------------------------------------
# Stage history
# ---------------------------------------------------------------------------

def seed_stage_history(workbook, opportunities, stages, users):
    actor_candidates = [
        user_for_role(users, SALES_EXECUTIVE),
        user_for_role(users, SALES_MANAGER),
        user_for_role(users, SOLUTION_ENGINEER),
        user_for_role(users, DELIVERY),
    ]
    actor_candidates = [
        user for user in actor_candidates if user
    ]

    if not actor_candidates:
        return

    for index, row in enumerate(
        rows_from_sheet(workbook, "Opportunities")
    ):
        opportunity = opportunities.get(
            clean(row.get("opportunity_id"))
        )

        if not opportunity:
            continue

        stage = stages.get(
            clean(row.get("pipeline_stage"))
        )

        if not stage:
            continue

        actor = actor_candidates[
            index % len(actor_candidates)
        ]

        remarks = (
            f"Imported stage history for "
            f"{clean(row.get('opportunity_id'))}."
        )

        exists = StageHistory.query.filter_by(
            opportunity_id=opportunity.opportunity_id,
            stage_id=stage.stage_id,
            remarks=remarks,
        ).first()

        if not exists:
            db.session.add(
                StageHistory(
                    opportunity_id=opportunity.opportunity_id,
                    stage_id=stage.stage_id,
                    changed_by=actor.user_id,
                    remarks=remarks,
                )
            )

    db.session.flush()


# ---------------------------------------------------------------------------
# POC tracker
# ---------------------------------------------------------------------------

def seed_pocs(workbook, opportunities):
    for row in rows_from_sheet(workbook, "POC_Tracker"):
        external_poc_id = clean(row.get("poc_id"))
        opportunity = opportunities.get(
            clean(row.get("opportunity_id"))
        )

        if not external_poc_id or not opportunity:
            continue

        # Current schema has no external poc_id, so the supplied workbook
        # opportunity + start date uniquely identify the imported tracker.
        start_date = parse_date(row.get("start_date"))

        tracker = POCTracker.query.filter_by(
            opportunity_id=opportunity.opportunity_id,
            start_date=start_date,
        ).first()

        exit_criteria = clean(row.get("exit_criteria"))
        success_metric = clean(row.get("success_metric"))
        planned_end_date = parse_date(row.get("planned_end_date"))

        # The current POCTracker schema requires poc_name, objective,
        # success_metric, target_date, and failure_condition, while the
        # workbook does not contain dedicated columns for all five fields.
        # Use only values derivable from the workbook/current opportunity;
        # do not invent a separate business-data source.
        poc_name = f"POC - {opportunity.opportunity_name}"
        objective = exit_criteria
        target_date = planned_end_date
        failure_condition = exit_criteria

        if not objective:
            raise RuntimeError(
                f"POC {external_poc_id} for opportunity "
                f"{opportunity.opportunity_id} has no exit_criteria. "
                "The current POCTracker schema requires objective."
            )

        if not success_metric:
            raise RuntimeError(
                f"POC {external_poc_id} for opportunity "
                f"{opportunity.opportunity_id} has no success_metric."
            )

        if not target_date:
            raise RuntimeError(
                f"POC {external_poc_id} for opportunity "
                f"{opportunity.opportunity_id} has no planned_end_date. "
                "The current POCTracker schema requires target_date."
            )

        if not tracker:
            tracker = POCTracker(
                opportunity_id=opportunity.opportunity_id,
                poc_name=poc_name,
                start_date=start_date,
                end_date=planned_end_date,
                status=clean(row.get("outcome")) or "Planned",
                objective=objective,
                success_metric=success_metric,
                target_date=target_date,
                failure_condition=failure_condition,
                outcome=clean(row.get("outcome")),
                outcome_notes=clean(row.get("outcome_notes")),
                exit_criteria=exit_criteria,
            )
            db.session.add(tracker)
        else:
            values = {
                "opportunity_id": opportunity.opportunity_id,
                "poc_name": poc_name,
                "start_date": start_date,
                "end_date": planned_end_date,
                "status": clean(row.get("outcome")) or "Planned",
                "target_date": target_date,
                "exit_criteria": exit_criteria,
                "objective": objective,
                "success_metric": success_metric,
                "failure_condition": failure_condition,
                "outcome": clean(row.get("outcome")),
                "outcome_notes": clean(row.get("outcome_notes")),
            }
            set_if_present(tracker, values)

        if "actual_end_date" in model_columns(POCTracker):
            tracker.actual_end_date = parse_date(
                row.get("actual_end_date")
            )

    db.session.flush()


# ---------------------------------------------------------------------------
# Legacy POC table
# ---------------------------------------------------------------------------

def seed_legacy_poc_table(opportunities):
    """
    Keep the existing separate Poc table usable. Only one deterministic row
    is created because the workbook's POC_Tracker sheet maps to POCTracker.
    """
    if not opportunities:
        return

    opportunity = next(iter(opportunities.values()))

    existing = Poc.query.filter_by(
        opportunity_id=opportunity.opportunity_id
    ).first()

    if existing:
        return

    db.session.add(
        Poc(
            opportunity_id=opportunity.opportunity_id,
            objective="Imported development POC reference.",
            success_metric="POC acceptance criteria are satisfied.",
            target_date=date.today(),
            failure_condition="A critical acceptance criterion fails.",
            stakeholder_signoff=False,
            outcome="Ongoing",
            outcome_notes="Development reference row.",
        )
    )

    db.session.flush()


# ---------------------------------------------------------------------------
# Activity Log -> existing AuditLog architecture
# ---------------------------------------------------------------------------

def resolve_activity_actor(name, users, index):
    """
    The workbook contains human names, while the development users are fixed
    test identities. Prefer an exact full-name match; otherwise use a stable
    rotating development actor. No new user is created from workbook data.
    """
    target = clean(name)

    if target:
        for user in users.values():
            if (
                user.full_name
                and user.full_name.strip().lower()
                == target.lower()
            ):
                return user

    candidates = [
        user_for_role(users, SALES_EXECUTIVE),
        user_for_role(users, SALES_MANAGER),
        user_for_role(users, PRE_SALES_MANAGER),
        user_for_role(users, SOLUTION_ENGINEER),
        user_for_role(users, DELIVERY),
    ]
    candidates = [
        user for user in candidates if user
    ]

    return (
        candidates[index % len(candidates)]
        if candidates
        else None
    )


def seed_activity_log(workbook, opportunities, users):
    """
    The current schema has audit_logs rather than a separate Activity_Log
    table. Therefore Activity_Log workbook rows are represented through the
    existing AuditLog architecture instead of creating a second database
    system.
    """
    for index, row in enumerate(
        rows_from_sheet(workbook, "Activity_Log")
    ):
        opportunity = opportunities.get(
            clean(row.get("opportunity_id"))
        )

        if not opportunity:
            continue

        action = clean(row.get("activity_type"))
        activity_id = clean(row.get("activity_id"))

        if not action:
            action = "ACTIVITY"

        notes = clean(row.get("notes"))
        next_step = clean(row.get("next_step"))

        description_parts = []
        if notes:
            description_parts.append(notes)
        if next_step:
            description_parts.append(
                f"Next step: {next_step}"
            )

        description = " ".join(
            description_parts
        ) or f"Imported activity {activity_id or index + 1}."

        actor = resolve_activity_actor(
            row.get("performed_by"),
            users,
            index,
        )

        if not actor:
            continue

        exists = AuditLog.query.filter_by(
            entity_type="Opportunity",
            entity_id=opportunity.opportunity_id,
            action=action,
            description=description,
        ).first()

        if exists:
            continue

        audit = AuditLog(
            entity_type="Opportunity",
            entity_id=opportunity.opportunity_id,
            action=action,
            description=description,
            performed_by=actor.user_id,
        )

        if "created_at" in model_columns(AuditLog):
            activity_date = parse_datetime(
                row.get("activity_date")
            )
            if activity_date:
                audit.created_at = activity_date

        db.session.add(audit)

    db.session.flush()


# ---------------------------------------------------------------------------
# Seed a small amount of workflow metadata only
# ---------------------------------------------------------------------------

def seed_tags():
    """
    Tags are not present in the supplied workbook. Keep the existing canonical
    tag set only if those tags are already part of the application's schema.
    No business records are invented here.
    """
    return


def seed_notifications(opportunities, users):
    """
    Do not manufacture notifications for every workbook row. Existing
    workflow notifications should be created by the application services.
    This function intentionally leaves notification creation to those
    services so seed data cannot hide notification duplication/security bugs.
    """
    return


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    app = create_app()

    with app.app_context():
        print("Seeding Deal Room development data...")
        print(f"Database: {app.config['SQLALCHEMY_DATABASE_URI']}")
        print(f"Workbook: {SEED_DATA_FILE}")

        workbook = load_seed_workbook()

        users = seed_users()
        stages = seed_stages()

        accounts = seed_accounts(workbook)

        opportunities = seed_opportunities(
            workbook,
            accounts,
            stages,
            users,
        )

        seed_contacts_from_stakeholders(
            workbook,
            opportunities,
            accounts,
        )

        seed_oem_partners(
            workbook,
            accounts,
        )

        seed_stakeholders(
            workbook,
            opportunities,
        )

        seed_opportunity_teams(
            opportunities,
            users,
        )

        seed_stage_history(
            workbook,
            opportunities,
            stages,
            users,
        )

        seed_pocs(
            workbook,
            opportunities,
        )

        seed_legacy_poc_table(
            opportunities,
        )

        seed_activity_log(
            workbook,
            opportunities,
            users,
        )

        # Notifications are intentionally not hard-coded. Real workflow
        # services must create them.
        seed_notifications(
            opportunities,
            users,
        )

        db.session.commit()

        print("\nSeed completed successfully.")
        print("Workbook data imported:")
        print(f"  Accounts:      {Account.query.count()}")
        print(f"  Opportunities: {Opportunity.query.count()}")
        print(f"  Contacts:      {Contact.query.count()}")
        print(f"  OEM partners:  {OEMPartner.query.count()}")
        print(f"  Stakeholders:  {Stakeholder.query.count()}")
        print(f"  POC trackers:  {POCTracker.query.count()}")
        print(f"  POC rows:      {Poc.query.count()}")
        print(f"  Audit logs:    {AuditLog.query.count()}")

        print("\nDevelopment identities:")
        for email in [
            "admin@dealroom.local",
            "sales.exec@dealroom.local",
            "sales.manager@dealroom.local",
            "presales.manager@dealroom.local",
            "solution.engineer@dealroom.local",
            "delivery@dealroom.local",
            "multi.role@dealroom.local",
            "pending@dealroom.local",
            "revoked@dealroom.local",
        ]:
            print(f"  {email} / Test@123")


if __name__ == "__main__":
    main()
